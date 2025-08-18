from openpyxl import load_workbook
from openpyxl.styles import Font

wb=load_workbook("basic.xlsx")
ws=wb.active
# print(ws["A1"].value)
for row in ws.iter_rows(values_only=True):
    print(row)
ws["A1"].font=Font(bold=True)
ws["B1"].font=Font(bold=True)

wb.save("basic.xlsx")