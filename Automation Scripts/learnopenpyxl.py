from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title="Logs"
# ws["A1"]="Hello Excel"
ws.cell(row=1,column=1,value="Name")
ws.cell(row=1,column=2,value="Age")
ws.cell(row=2,column=1,value="Alice")
ws.cell(row=2,column=2,value=25)
ws.cell(row=3,column=1,value="bob")
ws.cell(row=3,column=2,value=30)
# ws["B1"]="I am manav"
wb.save("basic.xlsx")

for row in ws.iter_rows(min_row=1,max_col=2,values_only=True):
    for value in row:
        print(value,end=" ")
    print()